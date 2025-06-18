import os
import json
import shutil
import httpx
import tempfile
import arabic_reshaper
from bidi.algorithm import get_display
from docx import Document
from PyPDF2 import PdfReader
import openpyxl
import streamlit as st
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import tiktoken
from openai import OpenAI

# Token-based chunking
def split_into_token_chunks(text, max_tokens=2000):
    try:
        enc = tiktoken.get_encoding("cl100k_base")
    except:
        return [text]

    paragraphs = text.split("\n\n")
    chunks = []
    current_chunk = ""
    current_tokens = 0

    for para in paragraphs:
        tokens = enc.encode(para)
        if current_tokens + len(tokens) > max_tokens:
            chunks.append(current_chunk.strip())
            current_chunk = para
            current_tokens = len(tokens)
        else:
            current_chunk += "\n\n" + para
            current_tokens += len(tokens)

    if current_chunk.strip():
        chunks.append(current_chunk.strip())
    return chunks

# Extract text

def extract_text(file):
    ext = os.path.splitext(file.name)[1].lower()
    if ext == ".docx":
        doc = Document(file)
        return "\n".join(p.text for p in doc.paragraphs)
    elif ext == ".pdf":
        reader = PdfReader(file)
        return "\n".join(page.extract_text() for page in reader.pages if page.extract_text())
    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(file, data_only=True)
        return extract_from_excel(wb)
    else:
        raise ValueError("Unsupported file type.")

def extract_from_excel(wb):
    text_blocks = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text_blocks.append(f"Sheet: {sheet}\n" + "-"*50)
        block = []
        for i, row in enumerate(ws.iter_rows()):
            row_data = [str(cell.value).strip() for cell in row if cell.value is not None]
            if row_data:
                block.append(" | ".join(row_data))
            if (i + 1) % 10 == 0 and block:
                text_blocks.append("\n".join(block))
                text_blocks.append("=" * 50)
                block = []
        if block:
            text_blocks.append("\n".join(block))
            text_blocks.append("=" * 50)
    return "\n\n".join(text_blocks)

# Generate PDF

def generate_pdf(text, output_path):
    font_path = "Amiri-Regular.ttf"
    pdfmetrics.registerFont(TTFont("ArabicFont", font_path))
    doc = SimpleDocTemplate(
        output_path, pagesize=letter,
        rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72
    )
    styles = getSampleStyleSheet()
    style = ParagraphStyle(
        'Arabic', parent=styles['Normal'], fontName="ArabicFont",
        fontSize=12, leading=14, rightIndent=0, alignment=2, spaceAfter=10
    )
    paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
    story = []
    for i, paragraph in enumerate(paragraphs):
        reshaped = arabic_reshaper.reshape(paragraph)
        bidi_text = get_display(reshaped)
        story.append(Paragraph(bidi_text, style))
        story.append(Spacer(1, 12))
    doc.build(story)

# Translation

def translate_text(text, target_language, update_progress):
    chunks = split_into_token_chunks(text)
    translated = []
    endpoint = "https://models.github.ai/inference"
    model = "openai/gpt-4.1"
    fallback_model = "openai/gpt-4.1-mini"
    token_map = {
        model: st.secrets["GITHUB_TOKEN_4_1"],
        fallback_model: st.secrets["GITHUB_TOKEN_4_1_MINI"]
    }
    if not token_map[model] or not token_map[fallback_model]:
        raise EnvironmentError("Model tokens not set in Streamlit secrets")

    for i, chunk in enumerate(chunks):
        try:
            client = OpenAI(base_url=endpoint, api_key=token_map[model])
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": f"You are a professional translator. Translate the following text to {target_language}. Preserve all layout and formatting (headings, indentation, table-like structures, line breaks). Do not add comments or explanations."},
                    {"role": "user", "content": chunk}
                ],
                temperature=0.3
            )
        except Exception as e:
            if "Rate limit" in str(e) or "no_access" in str(e):
                client = OpenAI(base_url=endpoint, api_key=token_map[fallback_model])
                response = client.chat.completions.create(
                    model=fallback_model,
                    messages=[
                        {"role": "system", "content": f"You are a professional translator. Translate the following text to {target_language}. Preserve all layout and formatting (headings, indentation, table-like structures, line breaks). Do not add comments or explanations."},
                        {"role": "user", "content": chunk}
                    ],
                    temperature=0.3
                )
            else:
                raise e
        translated.append(response.choices[0].message.content.strip())
        update_progress((i + 1) / len(chunks))
    return "\n\n".join(translated)

# Streamlit UI

st.set_page_config(page_title="Smart File Translator", layout="centered")
st.title("📁 Smart File Translator")

uploaded_file = st.file_uploader("Upload your DOCX, PDF, or XLSX file", type=["docx", "pdf", "xlsx"])
language = st.selectbox("Target Language", ["Arabic", "French", "English"], index=0)

if uploaded_file:
    if st.button("Translate"):
        try:
            with st.spinner("Reading file..."):
                raw_text = extract_text(uploaded_file)

            progress_bar = st.progress(0, text="Starting translation...")
            def update_progress(value):
                progress_bar.progress(value, text=f"Translating... {int(value*100)}%")

            translated = translate_text(raw_text, language, update_progress)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
                generate_pdf(translated, tmp_pdf.name)
                st.success("✅ Translation complete! Download below:")
                with open(tmp_pdf.name, "rb") as f:
                    st.download_button("📥 Download PDF", f, file_name="translated_output.pdf", mime="application/pdf")

        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

st.markdown("---")
st.caption("Developed by MAROUF ZOHRA © 2025")
